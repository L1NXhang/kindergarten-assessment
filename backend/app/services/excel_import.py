"""Excel导入服务

从Excel文件导入幼儿名单。
"""
import openpyxl
from app.extensions import db
from app.models.child import Child


class ExcelImportService:
    """Excel导入服务"""

    # 可能的姓名表头变体
    NAME_HEADERS = [
        '姓名', '学生姓名', '幼儿姓名', '幼儿名', '学生名',
        '孩子姓名', '儿童姓名', '名字', 'name', 'Name',
    ]

    # 可能的性别表头变体
    GENDER_HEADERS = [
        '性别', '性 别', 'gender', 'Gender',
    ]

    # 可能的出生日期表头变体
    BIRTH_DATE_HEADERS = [
        '出生日期', '出生年月', '生日', '出生', '出生时间',
        'birth_date', 'birthday', 'Birth',
    ]

    # 可能的备注表头变体
    NOTES_HEADERS = [
        '备注', '说明', '注释', 'notes', 'Notes', 'Remark',
    ]

    def import_children(self, file, class_id):
        """
        从Excel文件导入幼儿名单。

        参数:
            file: 文件对象（已通过验证）
            class_id: 班级ID

        返回:
            {
                'success': int,    # 成功导入数量
                'skipped': int,    # 跳过数量（重复）
                'errors': list,    # 错误信息列表
            }
        """
        result = {
            'success': 0,
            'skipped': 0,
            'errors': [],
        }

        try:
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            sheet = workbook.active

            # 自动识别表头
            header_map = self._detect_headers(sheet)

            if not header_map.get('name'):
                result['errors'].append('未找到姓名列，请检查表头')
                workbook.close()
                return result

            # 获取现有幼儿名单（用于去重）
            existing_children = set()
            existing = Child.query.filter_by(class_id=class_id).all()
            for child in existing:
                existing_children.add(child.name.strip())

            # 逐行读取数据
            row_num = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_num += 1

                if not row or all(cell is None for cell in row):
                    continue

                try:
                    name_col = header_map['name']
                    name = self._get_cell_value(row, name_col)

                    if not name:
                        continue

                    name = str(name).strip()

                    # 去重检查
                    if name in existing_children:
                        result['skipped'] += 1
                        continue

                    # 获取可选字段
                    gender = None
                    if 'gender' in header_map:
                        gender = self._get_cell_value(row, header_map['gender'])
                        if gender:
                            gender = str(gender).strip()
                            if gender not in ('男', '女'):
                                gender = None

                    birth_date = None
                    if 'birth_date' in header_map:
                        birth_date = self._get_cell_value(row, header_map['birth_date'])
                        birth_date = self._parse_date(birth_date)

                    notes = None
                    if 'notes' in header_map:
                        notes = self._get_cell_value(row, header_map['notes'])
                        if notes:
                            notes = str(notes).strip()

                    # 创建幼儿记录
                    child = Child(
                        name=name,
                        gender=gender,
                        birth_date=birth_date,
                        class_id=class_id,
                        notes=notes,
                    )
                    db.session.add(child)
                    existing_children.add(name)
                    result['success'] += 1

                except Exception as e:
                    result['errors'].append(f'第{row_num + 1}行: {str(e)}')

            db.session.commit()
            workbook.close()

        except Exception as e:
            db.session.rollback()
            result['errors'].append(f'文件处理失败: {str(e)}')

        return result

    def _detect_headers(self, sheet):
        """
        自动检测表头列映射。

        返回:
            {'name': int, 'gender': int, 'birth_date': int, 'notes': int}
        """
        header_map = {}

        # 读取第一行作为表头
        first_row = None
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            first_row = row
            break

        if not first_row:
            return header_map

        for col_idx, cell_value in enumerate(first_row):
            if cell_value is None:
                continue

            header_text = str(cell_value).strip().lower()

            if header_text in [h.lower() for h in self.NAME_HEADERS]:
                header_map['name'] = col_idx
            elif header_text in [h.lower() for h in self.GENDER_HEADERS]:
                header_map['gender'] = col_idx
            elif header_text in [h.lower() for h in self.BIRTH_DATE_HEADERS]:
                header_map['birth_date'] = col_idx
            elif header_text in [h.lower() for h in self.NOTES_HEADERS]:
                header_map['notes'] = col_idx

        # 如果没有找到表头，假设第一列是姓名
        if 'name' not in header_map and first_row:
            header_map['name'] = 0

        return header_map

    def _get_cell_value(self, row, col_idx):
        """安全获取单元格值"""
        if col_idx is not None and col_idx < len(row):
            return row[col_idx]
        return None

    def _parse_date(self, value):
        """解析日期值"""
        if value is None:
            return None

        from datetime import datetime

        # 如果是datetime对象
        if isinstance(value, datetime):
            return value.date()

        # 如果是date对象
        if hasattr(value, 'date'):
            return value.date()

        # 如果是字符串
        if isinstance(value, str):
            value = value.strip()
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%Y.%m.%d']:
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

        return None
