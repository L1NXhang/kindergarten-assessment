"""报告与导出路由"""
from datetime import datetime
from io import BytesIO
from flask import Blueprint, request, jsonify, send_file
from app.extensions import db, csrf
from app.models.child import Child
from app.models.class_group import ClassGroup
from app.models.anecdote import Anecdote
from app.models.anecdote_child import AnecdoteChild
from app.models.assessment import Assessment, AssessmentReport
from app.utils.response import success, error

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/reports/child/<int:child_id>/history', methods=['GET'])
@csrf.exempt
def get_child_report_history(child_id):
    """获取幼儿历史评估报告对比"""
    child = Child.query.get(child_id)
    if not child:
        return jsonify(error('幼儿不存在')), 404

    # 查询该幼儿的所有个人报告
    reports = (
        AssessmentReport.query
        .filter_by(child_id=child_id, report_type='individual')
        .join(Assessment, Assessment.id == AssessmentReport.assessment_id)
        .order_by(Assessment.period_start.desc())
        .all()
    )

    data = {
        'child': child.to_dict(include_class=True),
        'reports': [],
    }

    for report in reports:
        assessment = Assessment.query.get(report.assessment_id)
        item = report.to_dict()
        item['assessment_title'] = assessment.title if assessment else None
        item['period_start'] = assessment.period_start.isoformat() if assessment and assessment.period_start else None
        item['period_end'] = assessment.period_end.isoformat() if assessment and assessment.period_end else None
        data['reports'].append(item)

    return jsonify(success(data=data, message='获取幼儿历史报告成功'))


@reports_bp.route('/reports/class/<int:class_id>/history', methods=['GET'])
@csrf.exempt
def get_class_report_history(class_id):
    """获取班级历史评估报告对比"""
    class_group = ClassGroup.query.get(class_id)
    if not class_group:
        return jsonify(error('班级不存在')), 404

    # 查询该班级的所有评估
    assessments = (
        Assessment.query
        .filter_by(class_id=class_id)
        .order_by(Assessment.period_start.desc())
        .all()
    )

    data = {
        'class': class_group.to_dict(),
        'assessments': [],
    }

    for assessment in assessments:
        reports = AssessmentReport.query.filter_by(assessment_id=assessment.id).all()
        item = assessment.to_dict()
        item['reports'] = [r.to_dict() for r in reports]
        data['assessments'].append(item)

    return jsonify(success(data=data, message='获取班级历史报告成功'))


@reports_bp.route('/export/anecdotes', methods=['GET'])
@csrf.exempt
def export_anecdotes():
    """导出轶事记录为Excel"""
    class_id = request.args.get('class_id', type=int)
    export_format = request.args.get('format', 'xlsx')

    if export_format != 'xlsx':
        return jsonify(error('仅支持xlsx格式导出')), 400

    if not class_id:
        return jsonify(error('请指定班级ID')), 400

    class_group = ClassGroup.query.get(class_id)
    if not class_group:
        return jsonify(error('班级不存在')), 404

    # 查询轶事记录
    anecdotes = (
        Anecdote.query
        .filter_by(class_id=class_id)
        .order_by(Anecdote.observed_at.desc())
        .all()
    )

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{class_group.name}_轶事记录'

        # 表头样式
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(bold=True, size=12, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # 写入表头
        headers = ['序号', '观察日期', '幼儿姓名', '活动类型', '地点', '观察内容', '记录状态']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 写入数据
        for row_idx, anecdote in enumerate(anecdotes, 2):
            children_names = ', '.join([c.name for c in anecdote.children]) if anecdote.children else '未指定'
            date_str = anecdote.observed_at.strftime('%Y-%m-%d') if anecdote.observed_at else ''

            row_data = [
                row_idx - 1,
                date_str,
                children_names,
                anecdote.activity_type or '',
                anecdote.location or '',
                anecdote.content,
                anecdote.status or 'draft',
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 10

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'{class_group.name}_轶事记录_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )

    except ImportError:
        return jsonify(error('缺少openpyxl库，无法导出')), 500
    except Exception as e:
        return jsonify(error(f'导出失败: {str(e)}')), 500
